"""
Pipeline di ingestion documenti → DuckDB
Estrazione dati strutturati via Gemini API (primario) + Ollama (fallback)
Supporta: DOCX, PDF, XLSX, TXT, CSV
"""

import os
import re
import json
import hashlib
import logging
import time
from pathlib import Path
from typing import Optional

import duckdb
import requests
import openpyxl
from docx import Document

try:
    from pypdf import PdfReader
except ImportError:
    from PyPDF2 import PdfReader

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
log = logging.getLogger(__name__)

# ── Configurazione ──────────────────────────────────────────────────────────

DB_PATH       = os.path.expanduser("~/rag-pipeline/documenti.duckdb")
MAX_CHARS_LLM = 4000

# Gemini
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODEL   = "gemini-2.5-flash"
GEMINI_URL     = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"

# Ollama (fallback)
OLLAMA_URL    = "http://localhost:11434/api/generate"
OLLAMA_MODEL  = "qwen3:8b"
TIMEOUT_LLM   = 180

LLM_PROVIDER = "gemini" if GEMINI_API_KEY else "ollama"

ESTENSIONI_SUPPORTATE = {".docx", ".pdf", ".xlsx", ".xls", ".txt", ".csv", ".md"}


# ── Database ────────────────────────────────────────────────────────────────

def get_conn():
    return duckdb.connect(DB_PATH)


def init_db():
    conn = get_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS documenti (
            id          VARCHAR PRIMARY KEY,
            file_path   VARCHAR,
            file_name   VARCHAR,
            tipo        VARCHAR,
            testo       TEXT,
            metadata    JSON,
            hash_file   VARCHAR,
            data_import TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS fatture (
            id              VARCHAR PRIMARY KEY,
            file_path       VARCHAR,
            numero_fattura  VARCHAR,
            data_fattura    DATE,
            cliente         VARCHAR,
            totale          DOUBLE,
            voci            JSON,
            hash_file       VARCHAR,
            data_import     TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.close()
    log.info("Database inizializzato: %s [provider=%s]", DB_PATH, LLM_PROVIDER)


def already_imported(hash_val: str) -> bool:
    conn = get_conn()
    n = conn.execute(
        "SELECT COUNT(*) FROM documenti WHERE hash_file = ?", [hash_val]
    ).fetchone()[0]
    conn.close()
    return n > 0


def file_hash(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


# ── Estrazione testo grezzo ─────────────────────────────────────────────────

def estrai_docx(path: str) -> str:
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def estrai_pdf(path: str) -> str:
    reader = PdfReader(path)
    return "\n".join(
        p.extract_text() for p in reader.pages if p.extract_text()
    )


def estrai_xlsx(path: str) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    righe = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        righe.append(f"=== Foglio: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) if c is not None else "" for c in row]
            if any(v.strip() for v in vals):
                righe.append("\t".join(vals))
    return "\n".join(righe)


def estrai_testo(path: str) -> str:
    ext = Path(path).suffix.lower()
    try:
        if ext == ".docx":
            return estrai_docx(path)
        elif ext == ".pdf":
            return estrai_pdf(path)
        elif ext in (".xlsx", ".xls"):
            return estrai_xlsx(path)
        elif ext in (".txt", ".csv", ".md"):
            with open(path, encoding="utf-8", errors="replace") as f:
                return f.read()
        else:
            return ""
    except Exception as e:
        log.error("Errore estrazione testo %s: %s", path, e)
        return ""


# ── Prompt LLM ──────────────────────────────────────────────────────────────

PROMPT_CLASSIFICA = """Analizza questo documento e rispondi SOLO con un JSON valido, senza testo aggiuntivo, senza markdown, senza backtick.

Documento:
{testo}

Rispondi con questo schema JSON:
{{
  "tipo": "fattura" | "contratto" | "email" | "report" | "estratto_conto" | "altro",
  "è_documento_finanziario": true | false
}}"""

PROMPT_FATTURA = """Estrai i dati strutturati da questa fattura. Rispondi SOLO con un JSON valido, senza testo aggiuntivo, senza markdown, senza backtick.

Fattura:
{testo}

Per ogni voce della fattura, assegna una "categoria" con una o due parole in minuscolo che descriva il tipo di servizio/prodotto (esempi: "wordpress", "seo", "hosting", "grafica", "consulenza", "falegnameria", "marketing", ecc.). La categoria deve essere generica e riutilizzabile.

Rispondi con questo schema JSON esatto:
{{
  "numero_fattura": "stringa o null",
  "data_fattura": "YYYY-MM-DD o null",
  "cliente": "nome cliente o null",
  "fornitore": "nome fornitore o null",
  "totale": numero_decimale_o_null,
  "imponibile": numero_decimale_o_null,
  "iva": numero_decimale_o_null,
  "valuta": "EUR",
  "voci": [
    {{
      "descrizione": "descrizione originale della voce",
      "importo": numero,
      "categoria": "categoria_breve"
    }}
  ]
}}"""

PROMPT_GENERICO = """Estrai le informazioni principali da questo documento. Rispondi SOLO con un JSON valido, senza testo aggiuntivo, senza markdown, senza backtick.

Documento:
{testo}

Rispondi con questo schema JSON:
{{
  "titolo": "titolo o descrizione breve del documento",
  "data": "YYYY-MM-DD o null",
  "soggetti": ["lista di persone/aziende coinvolte"],
  "importi": [numero],
  "parole_chiave": ["max 5 parole chiave"],
  "sommario": "riassunto in 2 righe"
}}"""


# ── LLM: Gemini ─────────────────────────────────────────────────────────────

def chiama_gemini(prompt: str, retries: int = 2) -> Optional[str]:
    headers = {"Content-Type": "application/json"}
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0,
            "maxOutputTokens": 2048,
        }
    }
    for tentativo in range(retries + 1):
        try:
            url = f"{GEMINI_URL}?key={GEMINI_API_KEY}"
            resp = requests.post(url, headers=headers, json=payload, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            result = data["candidates"][0]["content"]["parts"][0]["text"].strip()
            return result if result else None
        except Exception as e:
            if tentativo < retries:
                log.warning("Retry Gemini %d: %s", tentativo + 1, e)
                time.sleep(1)
            else:
                log.error("Gemini non raggiungibile: %s", e)
                return None


# ── LLM: Ollama ─────────────────────────────────────────────────────────────

def chiama_ollama(prompt: str, retries: int = 2) -> Optional[str]:
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "options": {
            "temperature": 0,
            "num_predict": 2048,
            "num_ctx": 4096,
        }
    }
    for tentativo in range(retries + 1):
        try:
            resp = requests.post(OLLAMA_URL, json=payload, timeout=(10, TIMEOUT_LLM))
            resp.raise_for_status()
            raw = resp.json()
            result = raw.get("response", "").strip()
            if not result:
                result = raw.get("thinking", "").strip()
            return result if result else None
        except Exception as e:
            if tentativo < retries:
                log.warning("Retry Ollama %d: %s", tentativo + 1, e)
                time.sleep(2)
            else:
                log.error("Ollama non raggiungibile: %s", e)
                return None


# ── LLM unificato ───────────────────────────────────────────────────────────

def chiama_llm(prompt: str) -> Optional[str]:
    if LLM_PROVIDER == "gemini":
        result = chiama_gemini(prompt)
        if result:
            return result
        log.warning("Gemini fallito, uso Ollama come fallback")
    return chiama_ollama(prompt)


def pulisci_json(testo: str) -> Optional[dict]:
    if not testo:
        return None
    testo = re.sub(r"```(?:json)?", "", testo).strip().strip("`").strip()
    m = re.search(r"\{.*\}", testo, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass
    try:
        return json.loads(testo)
    except json.JSONDecodeError:
        log.warning("JSON non valido: %s", testo[:200])
        return None


# ── Classificazione e estrazione ────────────────────────────────────────────

def classifica_documento(testo: str) -> dict:
    prompt = PROMPT_CLASSIFICA.format(testo=testo[:MAX_CHARS_LLM])
    risposta = chiama_llm(prompt)
    risultato = pulisci_json(risposta)
    if risultato:
        return risultato
    testo_lower = testo.lower()
    if any(k in testo_lower for k in ["fattura", "invoice", "parcella"]):
        return {"tipo": "fattura", "è_documento_finanziario": True}
    return {"tipo": "altro", "è_documento_finanziario": False}


def estrai_dati_fattura_llm(testo: str) -> Optional[dict]:
    prompt = PROMPT_FATTURA.format(testo=testo[:MAX_CHARS_LLM])
    risposta = chiama_llm(prompt)
    return pulisci_json(risposta)


def estrai_dati_generico_llm(testo: str) -> Optional[dict]:
    prompt = PROMPT_GENERICO.format(testo=testo[:MAX_CHARS_LLM])
    risposta = chiama_llm(prompt)
    return pulisci_json(risposta)


# ── Import singolo file ─────────────────────────────────────────────────────

def importa_file(path: str) -> dict:
    path = os.path.abspath(path)
    file_name = os.path.basename(path)
    ext = Path(path).suffix.lower()
    hash_val = file_hash(path)

    risultato = {"file": file_name, "status": None, "tipo": None}

    if already_imported(hash_val):
        risultato["status"] = "già_importato"
        return risultato

    testo = estrai_testo(path)
    if not testo.strip():
        risultato["status"] = "testo_vuoto"
        return risultato

    classificazione = classifica_documento(testo)
    tipo = classificazione.get("tipo", "altro")

    metadata = {
        "estensione": ext,
        "dimensione": os.path.getsize(path),
        "tipo_rilevato": tipo,
        "classificazione": classificazione
    }

    conn = get_conn()
    try:
        conn.execute("""
            INSERT OR REPLACE INTO documenti
              (id, file_path, file_name, tipo, testo, metadata, hash_file)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, [hash_val, path, file_name, tipo, testo,
              json.dumps(metadata), hash_val])

        risultato["tipo"] = tipo

        if tipo == "fattura":
            dati = estrai_dati_fattura_llm(testo)
            if dati and dati.get("totale") is not None:
                try:
                    totale = float(str(dati["totale"]).replace(",", "."))
                except (ValueError, TypeError):
                    totale = None

                conn.execute("""
                    INSERT OR REPLACE INTO fatture
                      (id, file_path, numero_fattura, data_fattura,
                       cliente, totale, voci, hash_file)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, [
                    hash_val, path,
                    dati.get("numero_fattura"),
                    dati.get("data_fattura"),
                    dati.get("cliente") or dati.get("fornitore"),
                    totale,
                    json.dumps(dati.get("voci", [])),
                    hash_val
                ])
                risultato["totale"] = totale

        conn.commit()
        risultato["status"] = "importato"

    except Exception as e:
        log.error("Errore DB per %s: %s", file_name, e)
        risultato["status"] = f"errore: {e}"
    finally:
        conn.close()

    return risultato


# ── Import cartella ─────────────────────────────────────────────────────────

def importa_cartella(cartella: str, ricorsivo: bool = True) -> dict:
    cartella = os.path.abspath(cartella)
    pattern = "**/*" if ricorsivo else "*"
    files = [
        str(p) for p in Path(cartella).glob(pattern)
        if p.is_file() and p.suffix.lower() in ESTENSIONI_SUPPORTATE
    ]

    log.info("Trovati %d file in %s [provider=%s]", len(files), cartella, LLM_PROVIDER)
    stats = {"totale": len(files), "importati": 0,
             "già_importati": 0, "errori": 0}

    for i, fp in enumerate(files, 1):
        log.info("[%d/%d] %s", i, len(files), os.path.basename(fp))
        res = importa_file(fp)
        if res["status"] == "importato":
            stats["importati"] += 1
            extra = f" → {res['totale']:.2f} EUR" if res.get("totale") else ""
            log.info("  ✓ %s (%s)%s", res["file"], res["tipo"], extra)
        elif res["status"] == "già_importato":
            stats["già_importati"] += 1
        else:
            stats["errori"] += 1
            log.warning("  ✗ %s → %s", res["file"], res["status"])

    return stats


# ── Query helper ────────────────────────────────────────────────────────────

def query(sql: str):
    conn = get_conn()
    try:
        return conn.execute(sql).df()
    finally:
        conn.close()


# ── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    init_db()

    if len(sys.argv) < 2:
        print("Uso: GEMINI_API_KEY=... python ingestion.py <cartella> [--reset]")
        sys.exit(1)

    percorso = sys.argv[1]

    if "--reset" in sys.argv:
        conn = get_conn()
        conn.execute("DELETE FROM documenti")
        conn.execute("DELETE FROM fatture")
        conn.commit()
        conn.close()
        log.info("Database svuotato.")

    t_start = time.time()

    if os.path.isdir(percorso):
        stats = importa_cartella(percorso)
        durata = time.time() - t_start
        print(f"\n{'='*50}")
        print(f"Importazione completata in {durata:.0f}s [{LLM_PROVIDER}]:")
        print(f"  Processati:    {stats['totale']}")
        print(f"  Importati:     {stats['importati']}")
        print(f"  Già presenti:  {stats['già_importati']}")
        print(f"  Errori:        {stats['errori']}")
    else:
        res = importa_file(percorso)
        print(f"\nRisultato: {res}")

    # Riepilogo
    print(f"\n{'='*50}")
    conn = get_conn()
    n_doc  = conn.execute("SELECT COUNT(*) FROM documenti").fetchone()[0]
    n_fatt = conn.execute("SELECT COUNT(*) FROM fatture").fetchone()[0]
    print(f"Documenti totali: {n_doc}")
    print(f"Fatture:          {n_fatt}")
    if n_fatt > 0:
        totale = conn.execute(
            "SELECT SUM(totale) FROM fatture WHERE totale IS NOT NULL"
        ).fetchone()[0]
        if totale:
            print(f"Fatturato totale: {totale:,.2f} EUR")

        print(f"\nCategorie trovate nelle voci:")
        try:
            df = conn.execute("""
                SELECT DISTINCT json_extract_string(voce, '$.categoria') as categoria,
                       COUNT(*) as occorrenze
                FROM fatture, UNNEST(CAST(voci AS JSON[])) AS t(voce)
                WHERE json_extract_string(voce, '$.categoria') IS NOT NULL
                GROUP BY categoria
                ORDER BY occorrenze DESC
            """).df()
            print(df.to_string(index=False))
        except Exception as e:
            log.warning("Impossibile mostrare categorie: %s", e)
    conn.close()
